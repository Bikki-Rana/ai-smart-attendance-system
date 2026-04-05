"""
Export utilities: CSV and Excel export of attendance data
"""
import csv
import io
import os
from datetime import datetime
from typing import List, Dict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def export_to_csv(records: List[Dict], filename: str = None) -> str:
    """Export records to CSV, returns file path."""
    if not filename:
        filename = f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    export_dir = "./exports"
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)

    if not records:
        with open(filepath, "w", newline="") as f:
            f.write("No records found\n")
        return filepath

    fieldnames = list(records[0].keys())
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    return filepath


def export_to_excel(records: List[Dict], filename: str = None, title: str = "Attendance Report") -> str:
    """Export records to styled Excel file, returns file path."""
    if not filename:
        filename = f"attendance_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    export_dir = "./exports"
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, filename)

    if not EXCEL_AVAILABLE:
        return export_to_csv(records, filename.replace(".xlsx", ".csv"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"SmartAttend — {title}"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1E3A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(italic=True, color="666666")

    if not records:
        wb.save(filepath)
        return filepath

    # Header row
    headers = list(records[0].keys())
    header_row = 4
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header.replace("_", " ").title())
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, record in enumerate(records, header_row + 1):
        fill_color = "F8FAFC" if row_idx % 2 == 0 else "FFFFFF"
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(record.get(key, "")))
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center")

    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    wb.save(filepath)
    return filepath
